# -*- coding:utf-8 -*-

import json
import os
import openpyxl
import requests
from itertools import islice

excel_file = 'data.xlsx'
if (os.path.isfile(excel_file)):
    pass
else:
    pass


def convert_str_dict(value):
    if value != None:
        if isinstance(value, str):
            value = value.replace('\n', '').replace('\b', '').replace(' ', '')
            try:
                if '{' in value:
                    value = json.loads(value)
            except Exception:
                print('format error!Only json data available!')
                print(value)
    # print(type(value),value)
    return value


wb = openpyxl.load_workbook(excel_file)
print(wb.sheetnames)
print(wb.active)
sheet = wb['testcase']
# print(sheet.max_column)
rows_num = sheet.max_row

# for lines in islice(sheet,1,None):
#     for cell in lines:
#         value = cell.value
#         value = convert_str_dict(value)
#         print(value)
request_timeout = 10
for row in range(2, rows_num):
    # print('11'.center(120, '-'))
    desception = sheet.cell(row, 3).value
    url = sheet.cell(row, 4).value + sheet.cell(row, 5).value
    req_method = sheet.cell(row, 6).value
    req_header = convert_str_dict(sheet.cell(row, 7).value)
    param_or_body = convert_str_dict(sheet.cell(row, 8).value)
    param_body_value = convert_str_dict(sheet.cell(row, 9).value)
    correlation = convert_str_dict(sheet.cell(row, 10).value)
    checkpoint_code = convert_str_dict(sheet.cell(row, 12).value)
    assert_key = convert_str_dict(sheet.cell(row, 13).value)
    checkpoint_value = convert_str_dict(sheet.cell(row, 14).value)
    print('desception:', desception, '\n',
          'url:', url, '\n',
          'req_method:', req_method, '\n',
          'req_header:', req_header, '\n',
          'param_or_body:', param_or_body, '\n',
          'param_body_value:', param_body_value, '\n',
          'correlation:', correlation, '\n',
          'checkpoint_code:', checkpoint_code, '\n',
          'assert_key:', assert_key, '\n',
          'checkpoint_value:', checkpoint_value)
    print(type(correlation))
    for corr_key in correlation:
        if corr_key in req_header:
            req_header[corr_key] = correlation[corr_key]
            print(req_header[corr_key])
        elif corr_key in param_body_value:
            param_body_value[corr_key] = correlation[corr_key]
            print(param_body_value[corr_key])
        else:
            pass
    if req_method.lower() == 'get':
        pass
    elif req_method.lower() == 'post':
        pass
        # response = requests.post(url=url, headers=req_header, data=json.dumps(param_body_value), timeout=request_timeout)
    else:
        print('req_method ERROR')

